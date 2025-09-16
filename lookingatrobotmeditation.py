#!/usr/bin/env python3
import os
import re
import shutil
import unicodedata
from collections import Counter

import numpy as np
import pandas as pd
from openpyxl import load_workbook

# ========= Config =========
SOURCE_XLSX = "Shooter 2 Data.xlsx"
OUTPUT_XLSX = "Shooter 2 Proceed.xlsx"

# Sheets (0-based positions in workbook)
# 0: general (unused for writing here)
SHEET_PRE_NO   = 1  # pre-crisis, no meditation
SHEET_PRE_MED  = 2  # pre-crisis, meditation
SHEET_POST_NO  = 3  # post-crisis, no meditation
SHEET_POST_MED = 4  # post-crisis, meditation

# Write results to column index 1 (i.e., Excel column "B")
TARGET_COL_1BASED = 2

# CSV search folders (in order)
SEARCH_FOLDERS = [
    "shook",
    os.path.join("shook", "baseline"),
    "noshookmodified",
    os.path.join("noshookmodified", "baseline"),
]

# ========= Helpers =========
def normalize_text(s: str) -> str:
    """Trim, normalize unicode, collapse whitespace; preserve caller's case intent."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFC", str(s)).replace("\xa0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s

def norm_header_token(s: str) -> str:
    """Lowercase and remove spaces/underscores/dots/hyphens for robust header matching."""
    return re.sub(r"[ _.\-]+", "", normalize_text(s).lower())

def find_columns(header, want_keys):
    """
    Relaxed (case-insensitive) substring match on normalized header tokens.
    Returns dict {key -> col_idx or -1}.
    """
    normed = [norm_header_token(h) for h in header]
    out = {}
    for key in want_keys:
        k = norm_header_token(key)
        idx = -1
        for j, h in enumerate(normed):
            if k in h:
                idx = j
                break
        out[key] = idx
    return out

def index5(v) -> str:
    """First 5 chars of the Excel value; zero-pad if purely numeric and <5 digits."""
    if isinstance(v, (int, np.integer)):
        s = str(int(v)).zfill(5)
    elif isinstance(v, float) and float(v).is_integer():
        s = str(int(v)).zfill(5)
    else:
        s = str(v).strip()
        if s.isdigit() and len(s) < 5:
            s = s.zfill(5)
    return s[:5]

def load_csv_lenient(path: str) -> pd.DataFrame:
    return pd.read_csv(path, engine="python", on_bad_lines="skip", dtype=str)

def coerce_float(series: pd.Series) -> np.ndarray:
    return pd.to_numeric(series, errors="coerce").to_numpy()

def find_matching_csv(idx5: str):
    """Return (csv_path, folder_label) for first match, else (None, None)."""
    for folder in SEARCH_FOLDERS:
        if not os.path.isdir(folder):
            continue
        for name in os.listdir(folder):
            if name.lower().endswith(".csv") and name[:5] == idx5:
                return os.path.join(folder, name), folder
    return None, None

def crisis_time_from_robot_event(df: pd.DataFrame, time_col: str, robotevent_col: str):
    """
    Crisis-time detection (case-insensitive substr match within robotEvent column):
      1) first 'shook' row  -> t[row]
      2) else first '0.2 seconds' row -> t[row] + 0.229
      3) else -> None
    Returns (crisis_time or None, tag_src: 'shook'/'0.2 seconds'/None).
    """
    t = coerce_float(df[time_col])
    evt = df[robotevent_col].astype(str).str.lower()

    shook_mask = evt.str.contains("shook", regex=False, na=False)
    zero2_mask = evt.str.contains("0.2 seconds", regex=False, na=False)

    i_shook = int(np.flatnonzero(shook_mask.to_numpy())[0]) if shook_mask.any() else -1
    if i_shook != -1 and np.isfinite(t[i_shook]):
        return float(t[i_shook]), "shook"

    i_zero = int(np.flatnonzero(zero2_mask.to_numpy())[0]) if zero2_mask.any() else -1
    if i_zero != -1 and np.isfinite(t[i_zero]):
        return float(t[i_zero]) + 0.229, "0.2 seconds"

    return None, None

def meditation_toggles_from_event_column(df: pd.DataFrame, event_col_name: str):
    """
    Build dict: row_index -> 'enter' or 'exit' using the given event column.
    Tags (case-insensitive, substring in cell):
      - 'Entered Meditation Area' -> enter
      - 'Exited Meditation Area'  -> exit
    """
    toggles = {}
    ser = df[event_col_name].astype(str).str.lower()
    enter_mask = ser.str.contains("entered meditation area", regex=False, na=False)
    exit_mask  = ser.str.contains("exited meditation area",  regex=False, na=False)
    for i in np.flatnonzero(enter_mask.to_numpy()):
        toggles[int(i)] = "enter"
    for i in np.flatnonzero(exit_mask.to_numpy()):
        toggles[int(i)] = "exit"
    return toggles

def build_row_categories(t: np.ndarray, crisis_time: float, toggles: dict):
    """
    Assign each row to one of 4 categories (by row, not time-weighted):
      pre_no, pre_med, post_no, post_med.
    Toggle meditation state at the *row itself* (apply before classifying the row).
    Returns arrays of booleans for each category with length len(t).
    """
    n = len(t)
    in_med = False
    pre_no = np.zeros(n, dtype=bool)
    pre_med = np.zeros(n, dtype=bool)
    post_no = np.zeros(n, dtype=bool)
    post_med = np.zeros(n, dtype=bool)

    for i in range(n):
        act = toggles.get(i)
        if act == "enter":
            in_med = True
        elif act == "exit":
            in_med = False

        # pre vs post (exact equality goes to post)
        is_pre = float(t[i]) < float(crisis_time)
        if is_pre and not in_med:
            pre_no[i] = True
        elif is_pre and in_med:
            pre_med[i] = True
        elif (not is_pre) and not in_med:
            post_no[i] = True
        else:
            post_med[i] = True
    return pre_no, pre_med, post_no, post_med

def percent_looking_robot(df: pd.DataFrame, look_col_name: str, mask: np.ndarray) -> float:
    """Row-count %: (# rows in mask where lookingAt == 'robot') / (# rows in mask) * 100."""
    if mask is None or mask.sum() == 0:
        return float('nan')
    look = df[look_col_name].astype(str).str.strip().str.lower()
    num = (look.eq("robot") & mask).sum()
    den = int(mask.sum())
    return (num / den) * 100.0

# ========= Main =========
def main():
    # Duplicate workbook (don’t touch source)
    shutil.copyfile(SOURCE_XLSX, OUTPUT_XLSX)

    # Read indices from first sheet of source
    try:
        df_idx = pd.read_excel(SOURCE_XLSX, engine="openpyxl", sheet_name=0)
    except Exception as e:
        print(f"❌ Failed to read Excel '{SOURCE_XLSX}': {e}")
        return

    if df_idx.shape[0] == 0:
        print("❌ First sheet has no rows.")
        return

    indices = df_idx.iloc[:, 0].apply(index5)  # first column are indices

    # Open the duplicated workbook for writing with openpyxl
    try:
        wb = load_workbook(OUTPUT_XLSX)
    except Exception as e:
        print(f"❌ Failed to open '{OUTPUT_XLSX}' for writing: {e}")
        return

    # Ensure there are at least 5 sheets
    if len(wb.worksheets) < 5:
        print(f"❌ '{OUTPUT_XLSX}' must have at least 5 sheets (found {len(wb.worksheets)}).")
        return

    ws_pre_no   = wb.worksheets[SHEET_PRE_NO]
    ws_pre_med  = wb.worksheets[SHEET_PRE_MED]
    ws_post_no  = wb.worksheets[SHEET_POST_NO]
    ws_post_med = wb.worksheets[SHEET_POST_MED]

    # Skip tracking
    skip = Counter()
    ok = 0

    for row_idx, idx5 in enumerate(indices, start=2):  # Excel row (1-based). Row 1 is header.
        csv_path, folder = find_matching_csv(idx5)
        if not csv_path:
            skip["no_csv"] += 1
            # leave cells unchanged
            continue

        # Load CSV
        try:
            df = load_csv_lenient(csv_path)
        except Exception as e:
            print(f"{idx5}: SKIP CSV read error: {e} [{folder}]")
            skip["csv_read_error"] += 1
            continue
        if df.shape[0] == 0:
            print(f"{idx5}: SKIP empty CSV [{folder}]")
            skip["empty_csv"] += 1
            continue

        # Find required columns
        need = ["time", "robotevent", "roomevent", "event", "lookingat"]
        colmap = find_columns(list(df.columns), need)

        if colmap["time"] == -1 or colmap["robotevent"] == -1:
            print(f"{idx5}: SKIP missing Time/robotEvent columns [{folder}]")
            skip["missing_core_cols"] += 1
            continue

        time_col = df.columns[colmap["time"]]
        t = coerce_float(df[time_col])
        if len(t) == 0 or not np.any(np.isfinite(t)):
            print(f"{idx5}: SKIP no usable time series [{folder}]")
            skip["no_usable_time"] += 1
            continue

        # Crisis time
        robotevent_col = df.columns[colmap["robotevent"]]
        crisis_time, tag_src = crisis_time_from_robot_event(df, time_col, robotevent_col)
        if crisis_time is None:
            print(f"{idx5}: SKIP no 'shook' nor '0.2 seconds' tag [{folder}]")
            skip["no_crisis_tag"] += 1
            continue

        # lookingAt column
        if colmap["lookingat"] == -1:
            print(f"{idx5}: SKIP missing 'lookingAt' column [{folder}]")
            skip["missing_lookingAt"] += 1
            continue
        look_col = df.columns[colmap["lookingat"]]

        # Meditation toggles: prefer roomEvent; fallback Event; else no toggles
        toggles = {}
        if colmap["roomevent"] != -1:
            toggles = meditation_toggles_from_event_column(df, df.columns[colmap["roomevent"]])
        elif colmap["event"] != -1:
            toggles = meditation_toggles_from_event_column(df, df.columns[colmap["event"]])

        # Row category masks
        pre_no_m, pre_med_m, post_no_m, post_med_m = build_row_categories(t, crisis_time, toggles)

        # Compute % Looking At Robot by category (row-count based)
        pct_pre_no   = percent_looking_robot(df, look_col, pre_no_m)
        pct_pre_med  = percent_looking_robot(df, look_col, pre_med_m)
        pct_post_no  = percent_looking_robot(df, look_col, post_no_m)
        pct_post_med = percent_looking_robot(df, look_col, post_med_m)

        # Write to sheets (second column)
        ws_pre_no.cell(row=row_idx,   column=TARGET_COL_1BASED).value = None if np.isnan(pct_pre_no)   else float(pct_pre_no)
        ws_pre_med.cell(row=row_idx,  column=TARGET_COL_1BASED).value = None if np.isnan(pct_pre_med)  else float(pct_pre_med)
        ws_post_no.cell(row=row_idx,  column=TARGET_COL_1BASED).value = None if np.isnan(pct_post_no)  else float(pct_post_no)
        ws_post_med.cell(row=row_idx, column=TARGET_COL_1BASED).value = None if np.isnan(pct_post_med) else float(pct_post_med)

        ok += 1

    # Save workbook
    wb.save(OUTPUT_XLSX)

    # Summary
    print("\n===== SUMMARY =====")
    print(f"Indices found in sheet1: {len(indices)}")
    print(f"Processed successfully : {ok}")
    skipped = sum(skip.values())
    print(f"Skipped               : {skipped}")
    if skipped:
        for k, v in skip.items():
            print(f"  - {k}: {v}")
    print(f"\nWrote % Looking At Robot (row-count) to column B of sheets 2–5 in '{OUTPUT_XLSX}'.")
    print("Sheets: [2] pre_no, [3] pre_med, [4] post_no, [5] post_med")

if __name__ == "__main__":
    main()