#!/usr/bin/env python3
import os
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

INPUT_XLSX  = "Shooter 2 Data.xlsx"
OUTPUT_XLSX = "Shooter 2 Data Proceed.xlsx"
SEARCH_FOLDER = os.path.join("noshookmodified", "baseline")
ID_COL = 0  # first column (0-based)

# --- Highlight styles (tweak as you like) ---
USE_CELL_FILL = True
USE_FONT      = True

# Very visible magenta fill
FILL = PatternFill(patternType="solid", fgColor="FFFF00FF")  # Magenta

# Bold + red font so characters are obviously highlighted
FONT = Font(bold=True, color="FFFF0000")  # Red

def index5(v):
    """Zero-pad numerics to 5; otherwise string + zfill(5); then first 5 chars."""
    if isinstance(v, (int, float)) and float(v).is_integer():
        return str(int(v)).zfill(5)[:5]
    s = str(v).strip()
    return s.zfill(5)[:5]

def has_matching_csv(idx5: str) -> bool:
    if not os.path.isdir(SEARCH_FOLDER):
        return False
    for name in os.listdir(SEARCH_FOLDER):
        if name.lower().endswith(".csv") and name.startswith(idx5):
            return True
    return False

def main():
    if not os.path.isfile(INPUT_XLSX):
        print(f"❌ Input workbook not found: {INPUT_XLSX}")
        return

    # Duplicate workbook (don’t touch original)
    shutil.copyfile(INPUT_XLSX, OUTPUT_XLSX)
    print(f"📄 Duplicated: {INPUT_XLSX} → {OUTPUT_XLSX}")

    # Read indices from Sheet 1
    df = pd.read_excel(INPUT_XLSX, sheet_name=0, engine="openpyxl")
    if df.shape[1] <= ID_COL:
        print("❌ No index column (A) found.")
        return

    # Build (excel_row_number, idx5), header is row 1 → data starts at row 2
    indices = [(r + 2, index5(v)) for r, v in enumerate(df.iloc[:, ID_COL].tolist())]

    rows_to_highlight = []
    for xl_row, idx5 in indices:
        if has_matching_csv(idx5):
            rows_to_highlight.append((xl_row, idx5))

    # Apply highlight in the duplicate workbook
    wb = load_workbook(OUTPUT_XLSX)
    ws = wb[wb.sheetnames[0]]  # first sheet

    for r, idx5 in rows_to_highlight:
        cell = ws.cell(row=r, column=1)  # Column A
        if USE_CELL_FILL:
            cell.fill = FILL
        if USE_FONT:
            # Merge with any existing font settings
            cell.font = Font(
                name=cell.font.name,
                size=cell.font.size,
                bold=True if USE_FONT else cell.font.bold,
                color=FONT.color
            )

    wb.save(OUTPUT_XLSX)

    # Summary
    print(f"\n🔎 Searched: {SEARCH_FOLDER}")
    print(f"Total index rows read: {len(indices)}")
    print(f"Highlighted rows: {len(rows_to_highlight)}")
    if rows_to_highlight:
        ex = ", ".join([f"row {r} → {idx}" for r, idx in rows_to_highlight[:10]])
        print(f"Examples: {ex}")
    else:
        print("⚠️  No matches found. Confirm your CSVs start with the 5-char index.")

if __name__ == "__main__":
    main()