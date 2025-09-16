#!/usr/bin/env python3
import os
import re
import shutil
import unicodedata
import numpy as np
import pandas as pd
from openpyxl import load_workbook

# ========= Config =========
SOURCE_XLSX   = "Shooter 2 Data.xlsx"
OUTPUT_XLSX   = "Shooter 2 Proceed.xlsx"
ID_COL        = 0   # indices column on sheet 1 (0-based)
CRISIS_COL    = 3   # crisis time column on sheet 1 (0-based; Excel 4th column)

SEARCH_FOLDERS = [
    "shook",
    os.path.join("shook", "baseline"),
    "noshookmodified",
    os.path.join("noshookmodified", "baseline"),
]

# Sheet indices (1-based human; we convert to 0-based when indexing wb.worksheets)
SHEET_PRE_NO   = 2
SHEET_PRE_MED  = 3
SHEET_POST_NO  = 4
SHEET_POST_MED = 5

# Output columns (openpyxl is 1-based columns) → K,L,M,N
OUT_COL_MEAN = 11  # K
OUT_COL_SD   = 12  # L
OUT_COL_MAX  = 13  # M
OUT_COL_MIN  = 14  # N

# ========= Helpers =========
def normalize_text(s: str) -> str:
    if s is None:
        return ""
    s = unicodedata.normalize("NFC", str(s)).replace("\xa0", " ")
    return re.sub(r"\s+", " ", s).strip()

def norm_header_token(s: str) -> str:
    return re.sub(r"[ _.\-]+", "", normalize_text(s).lower())

def find_columns(header, want_keys):
    """
    Relaxed, case-insensitive substring match on normalized header tokens.
    Returns dict {key -> pandas column index or -1}.
    """
    normed = [norm_header_token(h) for h in header]
    out = {}
    for key in want_keys:
        k = norm_header_token(key)
        idx = -1
        for j, tok in enumerate(normed):
            if k in tok:
                idx = j
                break
        out[key] = idx
    return out

def index5(v) -> str:
    """Take first 5 chars; zero-pad if purely numeric and <5 digits."""
    if isinstance(v, (int, np.integer)):
        s = str(int(v)).zfill(5)
    elif isinstance(v, float) and float(v).is_integer():
        s = str(int(v)).zfill(5)
    else:
        s = str(v).strip()
        if s.isdigit() and len(s) < 5:
            s = s.zfill(5)
    return s[:5]

def find_matching_csv(idx5: str):
    """Return (csv_path, folder_label) or (None, None)."""
    for folder in SEARCH_FOLDERS:
        if not os.path.isdir(folder):
            continue
        for name in os.listdir(folder):
            if name.lower().endswith(".csv") and name.startswith(idx5):
                return os.path.join(folder, name), folder
    return None, None

def load_csv_lenient(path: str) -> pd.DataFrame:
    return pd.read_csv(path, engine="python", on_bad_lines="skip", dtype=str)

def coerce_float_series(ser: pd.Series) -> np.ndarray:
    return pd.to_numeric(ser, errors="coerce").to_numpy()

def meditation_toggles(df: pd.DataFrame, prefer_idx: int, fallback_idx: int):
    """
    Build {row_index -> 'enter'/'exit'} using roomEvent preferred, else Event.
    Tags (case-insensitive):
      'Entered Meditation Area' -> enter
      'Exited Meditation Area'  -> exit
    """
    src_idx = -1
    if prefer_idx != -1:
        src_idx = prefer_idx
    elif fallback_idx != -1:
        src_idx = fallback_idx
    else:
        return {}

    ser = df.iloc[:, src_idx].astype(str).str.lower()
    toggles = {}
    for i in np.flatnonzero(ser.str.contains("entered meditation area", regex=False, na=False).to_numpy()):
        toggles[int(i)] = "enter"
    for i in np.flatnonzero(ser.str.contains("exited meditation area", regex=False, na=False).to_numpy()):
        toggles[int(i)] = "exit"
    return toggles

def stats_or_nan(values):
    """Return (mean, sd(ddof=1), max, min) or all NaNs if fewer than 2 points."""
    n = len(values)
    if n < 2:
        return (np.nan, np.nan, np.nan, np.nan)
    arr = np.asarray(values, dtype=float)
    return (float(np.mean(arr)),
            float(np.std(arr, ddof=1)),
            float(np.max(arr)),
            float(np.min(arr)))

def write_stats(ws, row, stats_tuple):
    mean_v, sd_v, max_v, min_v = stats_tuple
    ws.cell(row=row, column=OUT_COL_MEAN, value=None if np.isnan(mean_v) else mean_v)
    ws.cell(row=row, column=OUT_COL_SD,   value=None if np.isnan(sd_v)   else sd_v)
    ws.cell(row=row, column=OUT_COL_MAX,  value=None if np.isnan(max_v)  else max_v)
    ws.cell(row=row, column=OUT_COL_MIN,  value=None if np.isnan(min_v)  else min_v)

# ========= Main =========
def main():
    # Duplicate workbook
    shutil.copyfile(SOURCE_XLSX, OUTPUT_XLSX)

    # Load sheet 1 for indices & crisis times
    try:
        df_idx = pd.read_excel(OUTPUT_XLSX, sheet_name=0, engine="openpyxl")
    except Exception as e:
        print(f"❌ Failed to read Excel '{OUTPUT_XLSX}': {e}")
        return

    if df_idx.shape[1] <= max(ID_COL, CRISIS_COL):
        print("❌ Sheet 1 must contain indices (col 1) and crisis time (col 4).")
        return

    indices = df_idx.iloc[:, ID_COL].apply(index5).tolist()
    crisis_times = pd.to_numeric(df_idx.iloc[:, CRISIS_COL], errors="coerce").to_numpy()

    # Open workbook for writing (sheets 2–5)
    try:
        wb = load_workbook(OUTPUT_XLSX)
    except Exception as e:
        print(f"❌ Failed to open workbook for writing: {e}")
        return

    if len(wb.worksheets) < 5:
        print("❌ Workbook must have at least 5 sheets (1 general + 4 category sheets).")
        return

    ws_pre_no   = wb.worksheets[SHEET_PRE_NO - 1]
    ws_pre_med  = wb.worksheets[SHEET_PRE_MED - 1]
    ws_post_no  = wb.worksheets[SHEET_POST_NO - 1]
    ws_post_med = wb.worksheets[SHEET_POST_MED - 1]

    skipped = 0

    for excel_row, (idx5, crisis_time) in enumerate(zip(indices, crisis_times), start=2):
        if not np.isfinite(crisis_time):
            print(f"{idx5}: SKIP (crisis time NaN on sheet 1)")
            skipped += 1
            continue

        csv_path, folder = find_matching_csv(idx5)
        if not csv_path:
            print(f"{idx5}: SKIP (no CSV found)")
            skipped += 1
            continue

        try:
            dfr = load_csv_lenient(csv_path)
        except Exception as e:
            print(f"{idx5}: SKIP (CSV read error: {e}) [{folder}]")
            skipped += 1
            continue
        if dfr.shape[0] == 0:
            print(f"{idx5}: SKIP (empty CSV) [{folder}]")
            skipped += 1
            continue

        # Resolve required columns
        need = [
            "time",
            "roomevent", "event",  # for meditation toggles (roomEvent preferred, else Event)
            "playervr.x", "playervr.y", "playervr.z",
            "robot.x",   "robot.y",   "robot.z",
        ]
        colmap = find_columns(list(dfr.columns), need)

        required = ["time", "playervr.x", "playervr.y", "playervr.z", "robot.x", "robot.y", "robot.z"]
        missing = [k for k in required if colmap[k] == -1]
        if missing:
            print(f"{idx5}: SKIP (missing columns: {', '.join(missing)}) [{folder}]")
            skipped += 1
            continue

        time_col = dfr.columns[colmap["time"]]
        times = coerce_float_series(dfr[time_col])

        # Meditation toggles (roomEvent preferred; else Event; else none)
        toggles = meditation_toggles(
            dfr,
            prefer_idx = colmap["roomevent"],
            fallback_idx = colmap["event"]
        )

        # Position columns
        pxc = dfr.columns[colmap["playervr.x"]]
        pyc = dfr.columns[colmap["playervr.y"]]
        pzc = dfr.columns[colmap["playervr.z"]]
        rxc = dfr.columns[colmap["robot.x"]]
        ryc = dfr.columns[colmap["robot.y"]]
        rzc = dfr.columns[colmap["robot.z"]]

        # Accumulators per category
        pre_no, pre_med, post_no, post_med = [], [], [], []
        in_med = False

        n = len(dfr)
        for i in range(n):
            # apply med toggle at this row
            act = toggles.get(i)
            if act == "enter":
                in_med = True
            elif act == "exit":
                in_med = False

            ti = times[i] if i < len(times) else np.nan
            if not np.isfinite(ti):
                continue

            # parse positions; keep -1 if present (treated as a real value)
            try:
                px = float(dfr.at[i, pxc]); py = float(dfr.at[i, pyc]); pz = float(dfr.at[i, pzc])
                rx = float(dfr.at[i, rxc]); ry = float(dfr.at[i, ryc]); rz = float(dfr.at[i, rzc])
            except Exception:
                continue

            dist = float(np.linalg.norm([px - rx, py - ry, pz - rz]))

            if ti <= crisis_time:
                (pre_med if in_med else pre_no).append(dist)
            else:
                (post_med if in_med else post_no).append(dist)

        # Compute stats
        pre_no_stats   = stats_or_nan(pre_no)
        pre_med_stats  = stats_or_nan(pre_med)
        post_no_stats  = stats_or_nan(post_no)
        post_med_stats = stats_or_nan(post_med)

        # Write to sheets 2–5 columns K–N
        write_stats(ws_pre_no,   excel_row, pre_no_stats)
        write_stats(ws_pre_med,  excel_row, pre_med_stats)
        write_stats(ws_post_no,  excel_row, post_no_stats)
        write_stats(ws_post_med, excel_row, post_med_stats)

        print(f"{idx5}: OK — pre_no(n={len(pre_no)}), pre_med(n={len(pre_med)}), "
              f"post_no(n={len(post_no)}), post_med(n={len(post_med)})")

    wb.save(OUTPUT_XLSX)
    print("\nDone. Wrote player–robot distance stats to sheets 2–5 (K–N) in", OUTPUT_XLSX)

if __name__ == "__main__":
    main()