#!/usr/bin/env python3
import os
import re
import shutil
import unicodedata
from collections import Counter

import numpy as np
import pandas as pd
from openpyxl import load_workbook

# ---------- Configuration ----------
SOURCE_XLSX = "Shooter 2 Data.xlsx"
OUTPUT_XLSX = "Shooter 2 Proceed.xlsx"
ID_COL = 0  # indices column in sheet 1 (0-based)

SEARCH_FOLDERS = [
    "shook",
    os.path.join("shook", "baseline"),
    "noshookmodified",
    os.path.join("noshookmodified", "baseline"),
]

# Category sheet order: [pre_no, pre_med, post_no, post_med]
SHEET_IDX_FOR_CAT = [1, 2, 3, 4]  # openpyxl 0-based list → sheets 2..5
WRITE_COLS_1BASED = (7, 8, 9, 10)  # G,H,I,J = Mean, SD, Min, Max

# ---------- Helpers ----------
def normalize_text(s: str) -> str:
    if s is None:
        return ""
    s = unicodedata.normalize("NFC", str(s)).replace("\xa0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s

def norm_header_token(s: str) -> str:
    return re.sub(r"[ _.\-]+", "", normalize_text(s).lower())

def find_columns(header, want_keys):
    normed = [norm_header_token(h) for h in header]
    out = {}
    for key in want_keys:
        k = norm_header_token(key)
        idx = -1
        for j, h in enumerate(normed):
            if k in h:
                idx = j
                break
        out[key] = idx
    return out

def index5(v) -> str:
    if isinstance(v, (int, np.integer)):
        s = str(int(v)).zfill(5)
    elif isinstance(v, float) and v.is_integer():
        s = str(int(v)).zfill(5)
    else:
        s = str(v).strip()
        if s.isdigit() and len(s) < 5:
            s = s.zfill(5)
    return s[:5]

def load_csv_lenient(path: str) -> pd.DataFrame:
    return pd.read_csv(path, engine="python", on_bad_lines="skip", dtype=str)

def coerce_float(series: pd.Series) -> np.ndarray:
    return pd.to_numeric(series, errors="coerce").to_numpy()

def find_matching_csv(idx5: str):
    for folder in SEARCH_FOLDERS:
        if not os.path.isdir(folder):
            continue
        for name in os.listdir(folder):
            if name.lower().endswith(".csv") and name[:5] == idx5:
                return os.path.join(folder, name), folder
    return None, None

# Crisis detection: robotEvent first, else Event; 'shook' first, then '0.2 seconds' (+0.229)
def crisis_time_from_tags(df: pd.DataFrame, time_col: str, tag_col: str):
    t = coerce_float(df[time_col])
    evt = df[tag_col].astype(str).map(lambda s: normalize_text(s).lower())

    m_shook = evt.str.contains("shook", regex=False, na=False)
    if m_shook.any():
        i = int(np.flatnonzero(m_shook.to_numpy())[0])
        if np.isfinite(t[i]):
            return float(t[i]), "shook"

    m_zero2 = evt.str.contains("0.2 seconds", regex=False, na=False)
    if m_zero2.any():
        i = int(np.flatnonzero(m_zero2.to_numpy())[0])
        if np.isfinite(t[i]):
            return float(t[i]) + 0.229, "0.2 seconds"

    return None, None

# Build in-meditation state from roomEvent (preferred) else Event
def build_in_meditation(df: pd.DataFrame, room_idx: int, event_idx: int, n_rows: int):
    used_idx = -1
    if room_idx != -1:
        used_idx = room_idx
    elif event_idx != -1:
        used_idx = event_idx
    else:
        return np.zeros(n_rows, dtype=bool), None

    colname = df.columns[used_idx]
    ser = df.iloc[:, used_idx].astype(str).map(lambda s: normalize_text(s).lower())

    enter_rows = set(np.flatnonzero(ser.str.contains("entered meditation area", regex=False, na=False)).tolist())
    exit_rows  = set(np.flatnonzero(ser.str.contains("exited meditation area",  regex=False, na=False)).tolist())

    in_med = np.zeros(n_rows, dtype=bool)
    state = False
    for i in range(n_rows):
        if i in enter_rows:
            state = True
        if i in exit_rows:
            state = False
        in_med[i] = state
    return in_med, colname

def categorize_row(ti: float, crisis_time: float, in_med: bool) -> str:
    if ti <= crisis_time:
        return "pre_med" if in_med else "pre_no"
    else:
        return "post_med" if in_med else "post_no"

def parse_pos3(x, y, z):
    try:
        fx, fy, fz = float(x), float(y), float(z)
    except Exception:
        return None
    # Treat -1 as invalid sample
    if fx == -1 or fy == -1 or fz == -1:
        return None
    return np.array([fx, fy, fz], dtype=float)

# ---------- Main ----------
def main():
    # Duplicate workbook
    shutil.copyfile(SOURCE_XLSX, OUTPUT_XLSX)

    # Read indices from sheet 1
    try:
        df_idx = pd.read_excel(SOURCE_XLSX, sheet_name=0, engine="openpyxl")
    except Exception as e:
        print(f"❌ Failed to read Excel '{SOURCE_XLSX}': {e}")
        return
    if df_idx.shape[1] <= ID_COL:
        print("❌ Excel does not have the indices column on sheet 1.")
        return
    indices = df_idx.iloc[:, ID_COL].apply(index5)

    # Open output workbook for writing to sheets 2–5
    try:
        wb = load_workbook(OUTPUT_XLSX)
    except Exception as e:
        print(f"❌ Failed to open '{OUTPUT_XLSX}' for writing: {e}")
        return
    if len(wb.worksheets) < 5:
        print("❌ Expected at least 5 sheets in the workbook (general + 4 category sheets).")
        return

    # Track issues
    skip_counts = Counter()
    ok_count = 0
    total_indices = 0
    not_found_csv = []
    missing_time = []
    missing_player_cols = []
    missing_crisis_cols = []
    no_crisis_tag = []
    csv_read_err = []
    insufficient_data = []

    def record_skip(reason_key: str, msg: str, bucket: list | None = None, idx: str | None = None):
        skip_counts[reason_key] += 1
        if bucket is not None and idx is not None:
            bucket.append(idx)
        print(msg)

    for excel_row_idx, idx_val in enumerate(indices, start=2):  # row 1 is header
        total_indices += 1
        idx5 = idx_val

        csv_path, folder = find_matching_csv(idx5)
        if not csv_path:
            record_skip("no_csv", f"{idx5}: SKIP (no CSV found in search folders)", not_found_csv, idx5)
            continue

        try:
            df = load_csv_lenient(csv_path)
        except Exception as e:
            record_skip("csv_read_error", f"{idx5}: SKIP (CSV read error: {e}) [{folder}]", csv_read_err, idx5)
            continue
        if df.shape[0] < 2:
            record_skip("too_few_rows", f"{idx5}: SKIP (CSV has <2 rows) [{folder}]")
            continue

        # Locate columns
        need = ["time", "playervr.x", "playervr.y", "playervr.z", "robotevent", "roomevent", "event"]
        colmap = find_columns(list(df.columns), need)

        # Time & Player pos required
        if colmap["time"] == -1:
            record_skip("missing_time_col", f"{idx5}: SKIP (missing 'Time' column) [{folder}]", missing_time, idx5)
            continue
        player_missing = [k for k in ("playervr.x","playervr.y","playervr.z") if colmap[k] == -1]
        if player_missing:
            record_skip("missing_player_cols",
                        f"{idx5}: SKIP (missing player columns: {', '.join(player_missing)}) [{folder}]",
                        missing_player_cols, idx5)
            continue

        time_col = df.columns[colmap["time"]]
        px_col = df.columns[colmap["playervr.x"]]
        py_col = df.columns[colmap["playervr.y"]]
        pz_col = df.columns[colmap["playervr.z"]]

        t = coerce_float(df[time_col])
        if not np.any(np.isfinite(t)):
            record_skip("no_usable_time", f"{idx5}: SKIP (no usable time values) [{folder}]")
            continue

        # Crisis detection: robotEvent preferred, else Event
        crisis_idx = colmap["robotevent"] if colmap["robotevent"] != -1 else colmap["event"]
        if crisis_idx == -1:
            record_skip("no_crisis_tag_col",
                        f"{idx5}: SKIP (missing both 'robotEvent' and 'Event' for crisis detection) [{folder}]",
                        missing_crisis_cols, idx5)
            continue
        crisis_col = df.columns[crisis_idx]
        crisis_time, tag_src = crisis_time_from_tags(df, time_col, crisis_col)
        if crisis_time is None:
            record_skip("no_crisis_tag",
                        f"{idx5}: SKIP (no 'shook' nor '0.2 seconds' tag found) [{folder}]",
                        no_crisis_tag, idx5)
            continue

        # Meditation state: roomEvent preferred, else Event
        in_med, used_med_col = build_in_meditation(df, colmap["roomevent"], colmap["event"], len(df))

        # Compute per-category velocity lists
        v_lists = {
            "pre_no": [],
            "pre_med": [],
            "post_no": [],
            "post_med": [],
        }

        # Iterate over segments (i -> i+1)
        px = df[px_col].to_numpy()
        py = df[py_col].to_numpy()
        pz = df[pz_col].to_numpy()
        n = len(df)

        for i in range(n - 1):
            if not (np.isfinite(t[i]) and np.isfinite(t[i+1])):
                continue
            dt = float(t[i+1]) - float(t[i])
            if dt <= 0:
                continue

            p1 = parse_pos3(px[i], py[i], pz[i])
            p2 = parse_pos3(px[i+1], py[i+1], pz[i+1])
            if p1 is None or p2 is None:
                continue

            speed = float(np.linalg.norm(p2 - p1) / dt)
            cat = categorize_row(float(t[i]), float(crisis_time), bool(in_med[i]))
            v_lists[cat].append(speed)

        # Compute stats and write to sheets 2..5, cols 7..10
        def stats(arr):
            if len(arr) == 0:
                return (np.nan, np.nan, np.nan, np.nan)
            mean = float(np.mean(arr))
            sd   = float(np.std(arr, ddof=1)) if len(arr) > 1 else float("nan")
            vmin = float(np.min(arr))
            vmax = float(np.max(arr))
            return (mean, sd, vmin, vmax)

        cats_order = ["pre_no", "pre_med", "post_no", "post_med"]
        wrote_any = False
        for sheet_idx, cat in zip(SHEET_IDX_FOR_CAT, cats_order):
            m, s, vmin, vmax = stats(v_lists[cat])
            ws = wb.worksheets[sheet_idx]
            ws.cell(row=excel_row_idx, column=WRITE_COLS_1BASED[0]).value = m
            ws.cell(row=excel_row_idx, column=WRITE_COLS_1BASED[1]).value = s
            ws.cell(row=excel_row_idx, column=WRITE_COLS_1BASED[2]).value = vmin
            ws.cell(row=excel_row_idx, column=WRITE_COLS_1BASED[3]).value = vmax
            wrote_any = wrote_any or (len(v_lists[cat]) > 0)

        if not wrote_any:
            insufficient_data.append(idx5)
            print(f"ℹ️ {idx5}: wrote NaNs (no valid velocity segments) [{folder}]")
        else:
            print(f"✅ {idx5} [{folder}] crisis via {tag_src or 'unknown'} ({crisis_col}) | "
                  f"segments: pre_no={len(v_lists['pre_no'])}, pre_med={len(v_lists['pre_med'])}, "
                  f"post_no={len(v_lists['post_no'])}, post_med={len(v_lists['post_med'])} "
                  f"(med from {used_med_col or 'none'})")

        ok_count += 1

    # Save workbook
    try:
        wb.save(OUTPUT_XLSX)
        print(f"\n💾 Wrote Player velocity stats to '{OUTPUT_XLSX}' (Sheets 2–5, cols G–J).")
    except Exception as e:
        print(f"❌ Failed to save '{OUTPUT_XLSX}': {e}")
        return

    # Summary
    print("\n===== SUMMARY =====")
    print(f"Total indices scanned: {total_indices}")
    print(f"Processed successfully: {ok_count}")
    total_skipped = sum(skip_counts.values())
    print(f"Skipped: {total_skipped}")
    if total_skipped:
        for reason, cnt in skip_counts.items():
            print(f"  - {reason}: {cnt}")

    if not_found_csv:
        print("\nMissing CSV:", ", ".join(not_found_csv))
    if missing_time:
        print("Missing 'Time' column:", ", ".join(missing_time))
    if missing_player_cols:
        print("Missing PlayerVR columns:", ", ".join(missing_player_cols))
    if missing_crisis_cols:
        print("Missing crisis tag columns ('robotEvent'/'Event'):", ", ".join(missing_crisis_cols))
    if no_crisis_tag:
        print("No crisis tag ('shook' or '0.2 seconds') found:", ", ".join(no_crisis_tag))
    if insufficient_data:
        print("No valid velocity segments (all categories NaN):", ", ".join(insufficient_data))

if __name__ == "__main__":
    main()