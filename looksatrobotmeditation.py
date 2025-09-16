#!/usr/bin/env python3
import os
import re
import shutil
import unicodedata
from collections import Counter

import numpy as np
import pandas as pd
from openpyxl import load_workbook

# ========== Config ==========
SOURCE_XLSX = "Shooter 2 Data.xlsx"
OUTPUT_XLSX = "Shooter 2 Proceed.xlsx"
ID_COL = 0  # first column in sheet 1 holds indices (same row order across sheets)

# Updated folders: shook + shook/baseline + noshookmodified + noshookmodified/baseline
SEARCH_FOLDERS = [
    "shook",
    os.path.join("shook", "baseline"),
    "noshookmodified",
    os.path.join("noshookmodified", "baseline"),
]

# Write targets: sheets 2..5 (1-based) -> indexes [1],[2],[3],[4] in openpyxl
# Order: pre_no, pre_med, post_no, post_med
SHEET_IDXS_FOR_CATEGORIES = [1, 2, 3, 4]
WRITE_COL_1BASED = 3  # column "C" (0-based index 2)

# ========== Text helpers ==========
def normalize_text(s: str) -> str:
    """Trim, normalize unicode, collapse spaces; preserve original chars (no .lower())."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFC", str(s)).replace("\xa0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s

def norm_header_token(s: str) -> str:
    """Normalize header for robust matching (case/space/dot/underscore/hyphen insensitive)."""
    return re.sub(r"[ _.\-]+", "", normalize_text(s).lower())

def find_columns(header, want_keys):
    """Relaxed (case-insensitive) substring match on normalized header tokens."""
    normed = [norm_header_token(h) for h in header]
    out = {}
    for key in want_keys:
        k = norm_header_token(key)
        idx = -1
        for j, h in enumerate(normed):
            if k in h:
                idx = j
                break
        out[key] = idx
    return out

def index5(v) -> str:
    """Take first 5 chars; zero-pad if purely numeric and <5 digits."""
    if isinstance(v, (int, np.integer)):
        s = str(int(v)).zfill(5)
    elif isinstance(v, float) and float(v).is_integer():
        s = str(int(v)).zfill(5)
    else:
        s = str(v).strip()
        if s.isdigit() and len(s) < 5:
            s = s.zfill(5)
    return s[:5]

# ========== CSV helpers ==========
def load_csv_lenient(path: str) -> pd.DataFrame:
    return pd.read_csv(path, engine="python", on_bad_lines="skip", dtype=str)

def coerce_float(series: pd.Series) -> np.ndarray:
    return pd.to_numeric(series, errors="coerce").to_numpy()

def find_matching_csv(idx5: str):
    """Return (csv_path, folder_label) for the first match, else (None, None)."""
    for folder in SEARCH_FOLDERS:
        if not os.path.isdir(folder):
            continue
        for name in os.listdir(folder):
            if name.lower().endswith(".csv") and name[:5] == idx5:
                return os.path.join(folder, name), folder
    return None, None

# ========== Crisis detection (shook first, then 0.2 seconds; robotEvent else Event) ==========
def crisis_time_from_tag_column(df: pd.DataFrame, time_col: str, tag_col_name: str):
    """
    Search (case-insensitive, trimmed, substring) in the provided tag column:
      1) first 'shook' row        -> t[row]
      2) else first '0.2 seconds' -> t[row] + 0.229
      3) else -> None
    Returns (crisis_time or None, 'shook'/'0.2 seconds'/None).
    """
    t = coerce_float(df[time_col])
    evt = df[tag_col_name].astype(str).map(lambda s: normalize_text(s).lower())

    m_shook = evt.str.contains("shook", regex=False, na=False)
    if m_shook.any():
        i = int(np.flatnonzero(m_shook.to_numpy())[0])
        if np.isfinite(t[i]):
            return float(t[i]), "shook"

    m_zero2 = evt.str_contains("0.2 seconds", regex=False, na=False) if hasattr(evt, "str_contains") else evt.str.contains("0.2 seconds", regex=False, na=False)
    if m_zero2.any():
        i = int(np.flatnonzero(m_zero2.to_numpy())[0])
        if np.isfinite(t[i]):
            return float(t[i]) + 0.229, "0.2 seconds"

    return None, None

# ========== Meditation toggles (roomEvent preferred, else Event) ==========
def build_meditation_state(df: pd.DataFrame, roomevent_idx: int, event_idx: int, n_rows: int):
    """
    Use roomEvent if present, otherwise Event.
    Tags (case-insensitive, substring):
      'entered meditation area' => enter (True)
      'exited meditation area'  => exit  (False)
    Returns (in_med: np.ndarray[bool], used_col: 'roomEvent'/'Event'/None)
    """
    in_med = np.zeros(n_rows, dtype=bool)
    used = None

    src_idx = -1
    if roomevent_idx != -1:
        src_idx = roomevent_idx
        used = "roomEvent"
    elif event_idx != -1:
        src_idx = event_idx
        used = "Event"
    else:
        return None, None

    ser = df.iloc[:, src_idx].astype(str).map(lambda s: normalize_text(s).lower())
    enters = set(np.flatnonzero(ser.str.contains("entered meditation area", regex=False, na=False)).tolist())
    exits  = set(np.flatnonzero(ser.str.contains("exited meditation area",  regex=False, na=False)).tolist())

    state = False
    for i in range(n_rows):
        if i in enters:
            state = True
        if i in exits:
            state = False
        in_med[i] = state

    return in_med, used

# ========== Episode counting per category ==========
def count_robot_looks_by_category(df: pd.DataFrame, time_col: str,
                                  in_med: np.ndarray, crisis_time: float,
                                  look_col_idx: int):
    """
    Count continuous 'lookingAt == robot' episodes, split into 4 categories by
    (pre/post crisis) × (in/out meditation).
    """
    t = coerce_float(df[time_col])
    look = df.iloc[:, look_col_idx].astype(str).map(lambda s: normalize_text(s).lower())

    cats = ["pre_no", "pre_med", "post_no", "post_med"]
    counts = {c: 0 for c in cats}
    rows_in_cat = {c: 0 for c in cats}

    def row_cat(i):
        if not np.isfinite(t[i]):
            return None
        if t[i] <= crisis_time:
            return "pre_med" if in_med[i] else "pre_no"
        else:
            return "post_med" if in_med[i] else "post_no"

    prev_on = False
    prev_cat = None

    n = len(t)
    for i in range(n):
        c = row_cat(i)
        if c is not None:
            rows_in_cat[c] += 1

        on = (look.iat[i] == "robot")
        if on and (not prev_on or c != prev_cat):
            if c is not None:
                counts[c] += 1

        prev_on = on
        prev_cat = c

    return counts, rows_in_cat

# ========== Main ==========
def main():
    # Duplicate workbook
    shutil.copyfile(SOURCE_XLSX, OUTPUT_XLSX)

    # Read indices from sheet 1
    try:
        df_idx = pd.read_excel(SOURCE_XLSX, sheet_name=0, engine="openpyxl")
    except Exception as e:
        print(f"❌ Failed to read Excel '{SOURCE_XLSX}': {e}")
        return
    if df_idx.shape[1] <= ID_COL:
        print("❌ Excel does not have the indices column on sheet 1.")
        return
    indices = df_idx.iloc[:, ID_COL].apply(index5)

    # Open output workbook
    try:
        wb = load_workbook(OUTPUT_XLSX)
    except Exception as e:
        print(f"❌ Failed to open '{OUTPUT_XLSX}' for writing: {e}")
        return
    if len(wb.worksheets) < 5:
        print("❌ Expected at least 5 sheets in the workbook.")
        return

    skip_counts = Counter()
    ok_count = 0
    total_indices = 0

    def record_skip(reason_key: str, msg: str):
        skip_counts[reason_key] += 1
        print(msg)

    for row_idx_excel, idx_val in enumerate(indices, start=2):  # header row is 1
        total_indices += 1
        idx5 = idx_val

        csv_path, folder = find_matching_csv(idx5)
        if not csv_path:
            record_skip("no_csv", f"{idx5}: SKIP (no CSV found) ")
            continue

        try:
            df = load_csv_lenient(csv_path)
        except Exception as e:
            record_skip("csv_read_error", f"{idx5}: SKIP (CSV read error: {e}) [{folder}]")
            continue
        if df.shape[0] == 0:
            record_skip("empty_csv", f"{idx5}: SKIP (empty CSV) [{folder}]")
            continue

        need = ["time", "robotevent", "roomevent", "event", "lookingat"]
        colmap = find_columns(list(df.columns), need)

        # time column required
        if colmap["time"] == -1:
            record_skip("missing_time_col", f"{idx5}: SKIP (missing 'Time' column) [{folder}]")
            continue
        time_col = df.columns[colmap["time"]]

        # lookingAt required
        if colmap["lookingat"] == -1:
            record_skip("missing_lookingAt_col", f"{idx5}: SKIP (missing 'lookingAt' column) [{folder}]")
            continue

        # ---- Crisis detection: ALWAYS prioritize 'shook' then '0.2 seconds',
        # using robotEvent if present, otherwise Event.
        crisis_col_idx = colmap["robotevent"] if colmap["robotevent"] != -1 else colmap["event"]
        if crisis_col_idx == -1:
            record_skip("missing_robot_or_event_for_crisis",
                        f"{idx5}: SKIP (missing both 'robotEvent' and 'Event' for crisis detection) [{folder}]")
            continue
        crisis_col_name = df.columns[crisis_col_idx]

        crisis_time, tag_src = crisis_time_from_tag_column(df, time_col, crisis_col_name)
        if crisis_time is None:
            record_skip("no_crisis_tag",
                        f"{idx5}: SKIP (no 'shook' nor '0.2 seconds' in {crisis_col_name}) [{folder}]")
            continue

        # ---- Meditation state: roomEvent preferred, else Event
        in_med, med_used_col = build_meditation_state(
            df,
            roomevent_idx=colmap["roomevent"],
            event_idx=colmap["event"],
            n_rows=len(df)
        )
        if in_med is None:
            record_skip("no_roomEvent_or_Event",
                        f"{idx5}: SKIP (missing both 'roomEvent' and 'Event' for meditation toggles) [{folder}]")
            continue

        # Count episodes by category
        counts, rows_in_cat = count_robot_looks_by_category(
            df, time_col, in_med, crisis_time, look_col_idx=colmap["lookingat"]
        )

        # Per-index console report
        pr = counts["pre_no"]; pm = counts["pre_med"]; r0 = rows_in_cat["pre_no"]; r1 = rows_in_cat["pre_med"]
        qr = counts["post_no"]; qm = counts["post_med"]; r2 = rows_in_cat["post_no"]; r3 = rows_in_cat["post_med"]
        print(
            f"{idx5} [{folder}] crisis={crisis_time:.3f}s (via {tag_src or 'unknown'} in {crisis_col_name}) | "
            f"pre_no={pr} (rows={r0}) pre_med={pm} (rows={r1}) | "
            f"post_no={qr} (rows={r2}) post_med={qm} (rows={r3})  med_tags={med_used_col}"
        )

        # Write to sheets 2..5, column C (blank if no rows for that category)
        values = [
            (counts["pre_no"],  rows_in_cat["pre_no"]),
            (counts["pre_med"], rows_in_cat["pre_med"]),
            (counts["post_no"], rows_in_cat["post_no"]),
            (counts["post_med"], rows_in_cat["post_med"]),
        ]
        try:
            for ws_idx, (cnt, nrows_cat) in zip(SHEET_IDXS_FOR_CATEGORIES, values):
                ws = wb.worksheets[ws_idx]
                cell = ws.cell(row=row_idx_excel, column=WRITE_COL_1BASED)
                cell.value = (None if nrows_cat == 0 else int(cnt))
            ok_count += 1
        except Exception as e:
            record_skip("excel_write_error", f"{idx5}: SKIP (failed writing to workbook: {e})")

    # Save workbook
    try:
        wb.save(OUTPUT_XLSX)
        print(f"\n💾 Wrote category look counts to '{OUTPUT_XLSX}' (Sheets 2–5, column C).")
    except Exception as e:
        print(f"❌ Failed to save '{OUTPUT_XLSX}': {e}")
        return

    # Summary
    print("\n===== SUMMARY =====")
    print(f"Total indices scanned: {total_indices}")
    print(f"Processed successfully: {ok_count}")
    total_skipped = sum(skip_counts.values())
    print(f"Skipped: {total_skipped}")
    if total_skipped:
        for reason, cnt in skip_counts.items():
            print(f"  - {reason}: {cnt}")

if __name__ == "__main__":
    main()