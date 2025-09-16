#!/usr/bin/env python3
import os
import re
import shutil
import unicodedata
from collections import Counter

import numpy as np
import pandas as pd
from openpyxl import load_workbook

# ========= Config =========
SOURCE_XLSX = "Shooter 2 Data.xlsx"        # input (sheet 1 has indices + crisis times)
OUTPUT_XLSX = "Shooter 2 TMP.xlsx"     # output (will be overwritten)
ID_COL = 0                                 # indices column on sheet 1 (0-based)
CRISIS_COL = 3                             # crisis time column on sheet 1 (0-based, i.e., 4th Excel column)

# Search folders in order
SEARCH_FOLDERS = [
    "shook",
    os.path.join("shook", "baseline"),
    "noshookmodified",
    os.path.join("noshookmodified", "baseline"),
]

# Sheet targets for the four categories (1-based positions in openpyxl)
# order = [pre_no, pre_med, post_no, post_med]
SHEET_IDX_FOR_CAT = [1, 2, 3, 4]      # sheets 2..5
WRITE_COL_1BASED = 4                  # write to Excel column D (1-based)

LOOK_TOKEN = "smoothfaced"            # <-- COUNT contiguous runs where LookingAt == "smoothfaced"

# ========= Helpers =========
def normalize_text(s: str) -> str:
    """Trim, NFC-normalize, collapse whitespace; keep original case for headers if needed."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFC", str(s)).replace("\xa0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s

def norm_header_token(s: str) -> str:
    """Lower/strip and remove spaces/._- for robust header substring matching."""
    return re.sub(r"[ _.\-]+", "", normalize_text(s).lower())

def find_columns(header, want_keys):
    """
    Relaxed (case-insensitive) substring match on normalized header tokens.
    Returns dict {key -> FIRST col_idx or -1}.
    """
    normed = [norm_header_token(h) for h in header]
    out = {}
    for key in want_keys:
        k = norm_header_token(key)
        idx = -1
        for j, h in enumerate(normed):
            if k in h:
                idx = j
                break
        out[key] = idx
    return out

def index5(v) -> str:
    """First 5 chars of the index cell; zero-pad if purely numeric and <5 digits."""
    if isinstance(v, (int, np.integer)):
        s = str(int(v)).zfill(5)
    elif isinstance(v, float) and float(v).is_integer():
        s = str(int(v)).zfill(5)
    else:
        s = str(v).strip()
        if s.isdigit() and len(s) < 5:
            s = s.zfill(5)
    return s[:5]

def load_csv_lenient(path: str) -> pd.DataFrame:
    return pd.read_csv(path, engine="python", on_bad_lines="skip", dtype=str)

def coerce_float(series: pd.Series) -> np.ndarray:
    return pd.to_numeric(series, errors="coerce").to_numpy()

def find_matching_csv(idx5: str):
    """Return (csv_path, folder_label) for the first match, else (None, None)."""
    for folder in SEARCH_FOLDERS:
        if not os.path.isdir(folder):
            continue
        for name in os.listdir(folder):
            if name.lower().endswith(".csv") and name.startswith(idx5):
                return os.path.join(folder, name), folder
    return None, None

def build_in_meditation(df: pd.DataFrame, room_idx: int, event_idx: int, n_rows: int):
    """
    Build per-row in-med flag from roomEvent preferred, else Event.
    Tags: 'entered meditation area' => enter; 'exited meditation area' => exit.
    Returns (in_med: np.ndarray[bool], used_col_name or None).
    """
    used_idx = -1
    used_name = None
    if room_idx != -1:
        used_idx = room_idx
        used_name = df.columns[room_idx]
    elif event_idx != -1:
        used_idx = event_idx
        used_name = df.columns[event_idx]
    else:
        return np.zeros(n_rows, dtype=bool), None

    ser = df.iloc[:, used_idx].astype(str).map(lambda s: normalize_text(s).lower())
    enter_rows = set(np.flatnonzero(ser.str.contains("entered meditation area", regex=False, na=False)).tolist())
    exit_rows  = set(np.flatnonzero(ser.str.contains("exited meditation area",  regex=False, na=False)).tolist())

    in_med = np.zeros(n_rows, dtype=bool)
    state = False
    for i in range(n_rows):
        if i in enter_rows:
            state = True
        if i in exit_rows:
            state = False
        in_med[i] = state
    return in_med, used_name

def categorize_row(ti: float, crisis_time: float, in_med: bool) -> str:
    """Return one of 'pre_no','pre_med','post_no','post_med' for the row time and med state."""
    if ti <= crisis_time:
        return "pre_med" if in_med else "pre_no"
    else:
        return "post_med" if in_med else "post_no"

# ========= Main =========
def main():
    # Duplicate workbook
    shutil.copyfile(SOURCE_XLSX, OUTPUT_XLSX)

    # Read indices & crisis times from sheet 1 (source)
    try:
        df_idx = pd.read_excel(SOURCE_XLSX, sheet_name=0, engine="openpyxl")
    except Exception as e:
        print(f"❌ Failed to read Excel '{SOURCE_XLSX}': {e}")
        return
    if df_idx.shape[1] <= max(ID_COL, CRISIS_COL):
        print("❌ Sheet 1 must contain indices (col 1) and crisis time (col 4).")
        return

    indices = df_idx.iloc[:, ID_COL].apply(index5)
    crisis_times = pd.to_numeric(df_idx.iloc[:, CRISIS_COL], errors="coerce").to_numpy()

    # Open duplicated workbook for writing
    try:
        wb = load_workbook(OUTPUT_XLSX)
    except Exception as e:
        print(f"❌ Failed to open '{OUTPUT_XLSX}' for writing: {e}")
        return
    if len(wb.worksheets) < 5:
        print("❌ Expected at least 5 sheets in the workbook.")
        return

    skip = Counter()
    ok = 0
    total = len(indices)

    for excel_row_idx, (idx5, crisis_time) in enumerate(zip(indices, crisis_times), start=2):  # row 1 is header
        if not np.isfinite(crisis_time):
            skip["crisis_time_nan"] += 1
            print(f"{idx5}: SKIP (crisis time NaN on sheet 1)")
            continue

        csv_path, folder = find_matching_csv(idx5)
        if not csv_path:
            skip["no_csv"] += 1
            print(f"{idx5}: SKIP (no CSV found)")
            continue

        # Load CSV
        try:
            df = load_csv_lenient(csv_path)
        except Exception as e:
            skip["csv_read_error"] += 1
            print(f"{idx5}: SKIP CSV read error: {e} [{folder}]")
            continue
        if df.shape[0] == 0:
            skip["empty_csv"] += 1
            print(f"{idx5}: SKIP empty CSV [{folder}]")
            continue

        # Resolve columns
        header = list(df.columns)
        colmap = find_columns(header, ["time", "roomevent", "event", "lookingat"])

        if colmap["time"] == -1 or colmap["lookingat"] == -1:
            if colmap["time"] == -1:
                skip["missing_time"] += 1
                print(f"{idx5}: SKIP missing 'Time' column [{folder}]")
            else:
                skip["missing_lookingat"] += 1
                print(f"{idx5}: SKIP missing 'LookingAt' column [{folder}]")
            continue

        time_col = header[colmap["time"]]
        look_col = header[colmap["lookingat"]]
        t = coerce_float(df[time_col])

        # Meditation state (roomEvent preferred else Event)
        in_med, used_med_col = build_in_meditation(df, colmap["roomevent"], colmap["event"], len(df))

        # Count contiguous “smoothfaced” looks in four buckets
        counts = {"pre_no": 0, "pre_med": 0, "post_no": 0, "post_med": 0}
        in_run = {"pre_no": False, "pre_med": False, "post_no": False, "post_med": False}
        prev_cat = None

        look_vals = df[look_col].astype(str).map(lambda s: normalize_text(s).lower()).to_numpy()

        for i in range(len(df)):
            ti = t[i]
            if not np.isfinite(ti):
                # break any ongoing run on invalid time
                prev_cat = None
                for k in in_run:
                    in_run[k] = False
                continue

            cat = categorize_row(float(ti), float(crisis_time), bool(in_med[i]))

            # If bucket changes, end any ongoing run
            if cat != prev_cat:
                for k in in_run:
                    in_run[k] = False
                prev_cat = cat

            on_face = (look_vals[i] == LOOK_TOKEN)  # <-- "smoothfaced"
            if on_face:
                if not in_run[cat]:
                    counts[cat] += 1
                    in_run[cat] = True
            else:
                in_run[cat] = False

        # Write counts to sheets 2..5 (column D)
        values = [counts["pre_no"], counts["pre_med"], counts["post_no"], counts["post_med"]]
        for sheet_1based, val in zip(SHEET_IDX_FOR_CAT, values):
            ws = wb.worksheets[sheet_1based]
            ws.cell(row=excel_row_idx, column=WRITE_COL_1BASED).value = int(val)

        ok += 1
        print(f"✅ {idx5} [{folder}] t_crisis={crisis_time:.3f}  "
              f"pre_no={counts['pre_no']}  pre_med={counts['pre_med']}  "
              f"post_no={counts['post_no']}  post_med={counts['post_med']}  "
              f"(med from {used_med_col or 'none'})")

    # Save workbook
    try:
        wb.save(OUTPUT_XLSX)
        print(f"\n💾 Wrote counts to sheets 2–5, column D in '{OUTPUT_XLSX}'.")
    except Exception as e:
        print(f"❌ Failed to save '{OUTPUT_XLSX}': {e}")
        return

    # Summary
    print("\n===== SUMMARY =====")
    print(f"Indices in sheet 1   : {total}")
    print(f"Processed successfully: {ok}")
    if skip:
        print("Skipped:")
        for k, v in skip.items():
            print(f"  - {k}: {v}")

if __name__ == "__main__":
    main()