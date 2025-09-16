#!/usr/bin/env python3
import os
import re
import shutil
import unicodedata
from collections import Counter

import numpy as np
import pandas as pd
from openpyxl import load_workbook

# ========= Config =========
SOURCE_XLSX = "Shooter 2 Data.xlsx"        # input (sheet 1 has indices)
OUTPUT_XLSX = "Shooter 2 Proceed.xlsx"     # output (will be overwritten)
ID_COL = 0                                 # indices column on sheet 1 (0-based)

SEARCH_FOLDERS = [
    "shook",
    os.path.join("shook", "baseline"),
    "noshookmodified",
    os.path.join("noshookmodified", "baseline"),
]

# Sheet targets for the four categories:
# order = [pre_no, pre_med, post_no, post_med]
SHEET_IDX_FOR_CAT = [1, 2, 3, 4]  # 1-based sheet numbers: sheets 2..5
WRITE_COL_1BASED = 6              # Column F

# ========= Helpers =========
def normalize_text(s: str) -> str:
    """Trim, NFC-normalize, collapse whitespace."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFC", str(s)).replace("\xa0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s

def norm_header_token(s: str) -> str:
    """Normalize header for robust substring matching."""
    return re.sub(r"[ _.\-]+", "", normalize_text(s).lower())

def find_columns(header, want_keys):
    """Relaxed (case-insensitive) substring match on normalized header tokens."""
    normed = [norm_header_token(h) for h in header]
    out = {}
    for key in want_keys:
        k = norm_header_token(key)
        idx = -1
        for j, h in enumerate(normed):
            if k in h:
                idx = j
                break
        out[key] = idx
    return out

def index5(v) -> str:
    """First 5 chars; zero-pad if purely numeric and <5 digits."""
    if isinstance(v, (int, np.integer)):
        s = str(int(v)).zfill(5)
    elif isinstance(v, float) and float(v).is_integer():
        s = str(int(v)).zfill(5)
    else:
        s = str(v).strip()
        if s.isdigit() and len(s) < 5:
            s = s.zfill(5)
    return s[:5]

def load_csv_lenient(path: str) -> pd.DataFrame:
    return pd.read_csv(path, engine="python", on_bad_lines="skip", dtype=str)

def coerce_float(series: pd.Series) -> np.ndarray:
    return pd.to_numeric(series, errors="coerce").to_numpy()

def find_matching_csv(idx5: str):
    """Return (csv_path, folder_label) for the first match (order = SEARCH_FOLDERS), else (None, None)."""
    for folder in SEARCH_FOLDERS:
        if not os.path.isdir(folder):
            continue
        for name in os.listdir(folder):
            if name.lower().endswith(".csv") and name[:5] == idx5:
                return os.path.join(folder, name), folder
    return None, None

# ---- Crisis detection: search robotEvent first, else Event; 'shook' first, then '0.2 seconds'
def crisis_time_from_tags(df: pd.DataFrame, time_col: str, tag_col: str):
    """
    Case-insensitive, trimmed, substring search within tag_col:
      1) first 'shook' row        -> t[row]
      2) else first '0.2 seconds' -> t[row] + 0.229
      3) else -> (None, None)
    """
    t = coerce_float(df[time_col])
    evt = df[tag_col].astype(str).map(lambda s: normalize_text(s).lower())

    m_shook = evt.str.contains("shook", regex=False, na=False)
    if m_shook.any():
        i = int(np.flatnonzero(m_shook.to_numpy())[0])
        if np.isfinite(t[i]):
            return float(t[i]), "shook"

    m_zero2 = evt.str.contains("0.2 seconds", regex=False, na=False)
    if m_zero2.any():
        i = int(np.flatnonzero(m_zero2.to_numpy())[0])
        if np.isfinite(t[i]):
            return float(t[i]) + 0.229, "0.2 seconds"

    return None, None

# ---- Build in-meditation boolean per row from roomEvent (preferred) or Event
def build_in_meditation(df: pd.DataFrame, room_idx: int, event_idx: int, n_rows: int):
    """
    Tags (case-insensitive substring):
      'entered meditation area' => enter
      'exited meditation area'  => exit
    Returns (in_med: np.ndarray[bool], used_col_name or None)
    """
    used_idx = -1
    used_name = None
    if room_idx != -1:
        used_idx = room_idx
        used_name = df.columns[room_idx]
    elif event_idx != -1:
        used_idx = event_idx
        used_name = df.columns[event_idx]
    else:
        return np.zeros(n_rows, dtype=bool), None

    ser = df.iloc[:, used_idx].astype(str).map(lambda s: normalize_text(s).lower())
    enter_rows = set(np.flatnonzero(ser.str.contains("entered meditation area", regex=False, na=False)).tolist())
    exit_rows  = set(np.flatnonzero(ser.str.contains("exited meditation area",  regex=False, na=False)).tolist())

    in_med = np.zeros(n_rows, dtype=bool)
    state = False
    for i in range(n_rows):
        if i in enter_rows:
            state = True
        if i in exit_rows:
            state = False
        in_med[i] = state
    return in_med, used_name

def categorize_row(ti: float, crisis_time: float, in_med: bool) -> str:
    """Return one of 'pre_no','pre_med','post_no','post_med' for the row time and med state."""
    if ti <= crisis_time:
        return "pre_med" if in_med else "pre_no"
    else:
        return "post_med" if in_med else "post_no"

# ========= Main =========
def main():
    # Duplicate workbook
    shutil.copyfile(SOURCE_XLSX, OUTPUT_XLSX)

    # Read indices from sheet 1
    try:
        df_idx = pd.read_excel(SOURCE_XLSX, sheet_name=0, engine="openpyxl")
    except Exception as e:
        print(f"❌ Failed to read Excel '{SOURCE_XLSX}': {e}")
        return
    if df_idx.shape[1] <= ID_COL:
        print("❌ Excel does not have the indices column on sheet 1.")
        return
    indices = df_idx.iloc[:, ID_COL].apply(index5)

    # Open output workbook
    try:
        wb = load_workbook(OUTPUT_XLSX)
    except Exception as e:
        print(f"❌ Failed to open '{OUTPUT_XLSX}' for writing: {e}")
        return
    if len(wb.worksheets) < 5:
        print("❌ Expected at least 5 sheets in the workbook.")
        return

    # Track skips/issues
    skip_counts = Counter()
    ok_count = 0
    total_indices = 0
    not_found_csv = []
    missing_lookingat = []
    missing_time = []
    missing_crisis_cols = []
    no_crisis_tag = []
    csv_read_err = []

    def record_skip(reason_key: str, msg: str, bucket: list | None = None, idx: str | None = None):
        skip_counts[reason_key] += 1
        if bucket is not None and idx is not None:
            bucket.append(idx)
        print(msg)

    for excel_row_idx, idx_val in enumerate(indices, start=2):  # row 1 = header in all sheets
        total_indices += 1
        idx5 = idx_val

        csv_path, folder = find_matching_csv(idx5)
        if not csv_path:
            record_skip("no_csv", f"{idx5}: SKIP (no CSV found in search folders)", not_found_csv, idx5)
            continue

        try:
            df = load_csv_lenient(csv_path)
        except Exception as e:
            record_skip("csv_read_error", f"{idx5}: SKIP (CSV read error: {e}) [{folder}]", csv_read_err, idx5)
            continue
        if df.shape[0] == 0:
            record_skip("empty_csv", f"{idx5}: SKIP (empty CSV) [{folder}]")
            continue

        # Locate columns
        need = ["time", "lookingat", "robotevent", "roomevent", "event"]
        colmap = find_columns(list(df.columns), need)

        # time + lookingAt required
        if colmap["time"] == -1:
            record_skip("missing_time_col", f"{idx5}: SKIP (missing 'Time' column) [{folder}]", missing_time, idx5)
            continue
        if colmap["lookingat"] == -1:
            record_skip("missing_lookingat_col", f"{idx5}: SKIP (missing 'LookingAt' column) [{folder}]", missing_lookingat, idx5)
            continue

        time_col = df.columns[colmap["time"]]
        look_col = df.columns[colmap["lookingat"]]

        t = coerce_float(df[time_col])
        if len(t) == 0 or not np.any(np.isfinite(t)):
            record_skip("no_usable_time", f"{idx5}: SKIP (no usable time values) [{folder}]")
            continue

        # Crisis detection uses robotEvent if present; else Event
        crisis_idx = colmap["robotevent"] if colmap["robotevent"] != -1 else colmap["event"]
        if crisis_idx == -1:
            record_skip("no_crisis_tag_col",
                        f"{idx5}: SKIP (missing both 'robotEvent' and 'Event' for crisis detection) [{folder}]",
                        missing_crisis_cols, idx5)
            continue
        crisis_col = df.columns[crisis_idx]
        crisis_time, tag_src = crisis_time_from_tags(df, time_col, crisis_col)
        if crisis_time is None:
            record_skip("no_crisis_tag",
                        f"{idx5}: SKIP (no 'shook' nor '0.2 seconds' tag found) [{folder}]",
                        no_crisis_tag, idx5)
            continue

        # Build in-meditation state (roomEvent preferred, else Event)
        in_med, used_med_col = build_in_meditation(df, colmap["roomevent"], colmap["event"], len(df))

        # Count contiguous runs of "signage" per category
        counts = {"pre_no": 0, "pre_med": 0, "post_no": 0, "post_med": 0}
        in_look = {"pre_no": False, "pre_med": False, "post_no": False, "post_med": False}
        prev_cat = None

        look_vals = df[look_col].astype(str).map(lambda s: normalize_text(s).lower()).to_numpy()
        n = len(df)

        for i in range(n):
            if not np.isfinite(t[i]):
                # break any ongoing run on invalid time
                prev_cat = None
                for k in in_look:
                    in_look[k] = False
                continue

            cat = categorize_row(float(t[i]), float(crisis_time), bool(in_med[i]))

            # Category change breaks ongoing look runs
            if cat != prev_cat:
                for k in in_look:
                    in_look[k] = False
                prev_cat = cat

            val = look_vals[i]
            # Keep original logic: startswith "signage" (case-insensitive)
            if isinstance(val, str) and val.startswith("signage"):
                if not in_look[cat]:
                    counts[cat] += 1
                    in_look[cat] = True
            else:
                in_look[cat] = False

        # Write counts to sheets 2..5 (column F)
        values = [counts["pre_no"], counts["pre_med"], counts["post_no"], counts["post_med"]]
        for sheet_idx, val in zip(SHEET_IDX_FOR_CAT, values):
            ws = wb.worksheets[sheet_idx]
            ws.cell(row=excel_row_idx, column=WRITE_COL_1BASED).value = int(val)

        ok_count += 1
        print(f"✅ {idx5} [{folder}] crisis via {tag_src or 'unknown'} ({crisis_col}) | "
              f"signage looks → pre_no={counts['pre_no']}  pre_med={counts['pre_med']}  "
              f"post_no={counts['post_no']}  post_med={counts['post_med']}  "
              f"(med from {used_med_col or 'none'})")

    # Save workbook
    try:
        wb.save(OUTPUT_XLSX)
        print(f"\n💾 Wrote signage look counts to '{OUTPUT_XLSX}' (Sheets 2–5, column F).")
    except Exception as e:
        print(f"❌ Failed to save '{OUTPUT_XLSX}': {e}")
        return

    # Summary
    print("\n===== SUMMARY =====")
    print(f"Total indices scanned: {total_indices}")
    print(f"Processed successfully: {ok_count}")
    total_skipped = sum(skip_counts.values())
    print(f"Skipped: {total_skipped}")
    if total_skipped:
        for reason, cnt in skip_counts.items():
            print(f"  - {reason}: {cnt}")

    # Problem buckets (optional lists)
    if not_found_csv:
        print("\nMissing CSV:", ", ".join(not_found_csv))
    if missing_lookingat:
        print("Missing 'LookingAt' column:", ", ".join(missing_lookingat))
    if missing_time:
        print("Missing 'Time' column:", ", ".join(missing_time))
    if missing_crisis_cols:
        print("Missing crisis tag columns ('robotEvent'/'Event'):", ", ".join(missing_crisis_cols))
    if no_crisis_tag:
        print("No crisis tag ('shook' or '0.2 seconds') found:", ", ".join(no_crisis_tag))

if __name__ == "__main__":
    main()